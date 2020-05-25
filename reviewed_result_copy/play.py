
from openpyxl import load_workbook, Workbook
from openpyxl.cell import Cell
from openpyxl.styles import PatternFill
from openpyxl.worksheet.worksheet import Worksheet

reviewed: Workbook = load_workbook(filename='reviewed.xlsm')
reviewedMenus: Worksheet = reviewed['Menus']

reviewedList = {}

for cells in reviewedMenus.rows:
    lineId: Cell = cells[0]
    corrections: Cell = cells[10]
    reviewedList[str(lineId.value)] = corrections.value

reviewed.close()

origin: Workbook = load_workbook(filename="origin.xlsm")
originMenus: Worksheet = origin['Menus']

for cells in originMenus.rows:
    lineId: Cell = cells[0]
    cell: Cell = cells[9]
    if lineId.value and lineId.value != "LineID" and lineId.value in reviewedList:
        print(lineId.value)
        print(cell.value)
        print(reviewedList[lineId.value])
        cell.value = reviewedList[lineId.value]
        cell.fill = PatternFill(fill_type="solid", start_color="ffff0000", end_color="ffff0000")
        print(cell.value)
        print("\n")


origin.save(filename="origin.xlsm")
origin.close()
