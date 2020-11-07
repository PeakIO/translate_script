import os
from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.worksheet import Worksheet

filename = "source"

wb: Workbook = load_workbook(filename=filename + '.xlsx')
ws: Worksheet = wb.active

az = "ABCDEFGHIJKLMNOPQRSTUVWXYZ"

res: dict = {}

answers = []
statements = []

for i, cells in enumerate(ws.iter_rows()):
    if i > 186:
        d = cells[az.index("D")].value
        e = cells[az.index("E")].value
        k = cells[az.index("K")].value
        v: str = cells[az.index("V")].value

        if v is not None and "Answers" in v:
            answers.append((d, e, k , v))
        else:
            statements.append((d, e, k , v))
                

result: list = []

names: list = ['Groot', 'Rocket', 'Starlord', 'Drax', 'Gamora', 'Glory', 'Mantis', 'Warlock', 'LadyHellbender']

def getName(input: str):
    res = input.split("_")
    name = ""
    if len(res[-1]) == 1:
        name = res[-2]
    else:
        name = res[-1]
    if name not in names:
        names.append(name)
    return name

def getIndex(input: str):
    for name in names:
        if name in input:
            return input.replace(name, "")

def joinName(input: str, name: str):
    return input[:18] + "Answers " + name + " " + input[18:]
    
wbb: Workbook = Workbook()
wss: Worksheet = wbb.active
last_ask = ""
last_answer = ""
for statement in statements:
    e1 = statement[1]
    v1 = statement[3]
    target_index = getIndex(e1)
    target_statement = joinName(v1, getName(e1))

    for answer in answers:
        e0: str = answer[1]
        v0: str = answer[3]
        if getIndex(e0) == target_index and target_statement == v0:
            if last_ask == statement[2]:
                wss.append(list(("","","","") + answer))
            elif last_answer == answer[2]:
                wss.append(list(statement + ("","","","")))
            else:
                wss.append(list(statement + answer))
            last_answer = answer[2]
            last_ask = statement[2]
wbb.save(filename="result.xlsx")
wbb.close()

wb.close()

print(names)