import os
from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.worksheet import Worksheet

filename = "08"

az = "ABCDEFGHIJKLMNOPQRSTUVWXYZ"
name_pos = az.index("B")
content_pos = az.index("A")

wb: Workbook = load_workbook(filename=filename + '.xlsx')
ws: Worksheet = wb.active

res: dict = {}

for i, cells in enumerate(ws.iter_rows()):
    if i != 0:
        person_name = cells[name_pos].value
        if person_name is not None:
            if person_name not in res:
                res[person_name] = []
            arr: list = res[person_name]
            arr.append(
                [
                    cells[content_pos].value,
                ],
            )

if not os.path.exists(filename):
    os.makedirs(filename)

for name, contents in res.items():
    wbb: Workbook = Workbook()
    wss: Worksheet = wbb.active
    count = 0
    for rows in contents:
        if rows[0] is not None:
            wss.append(rows)
            count = count + len(rows[0].strip().split())
    wbb.save(filename="./{path}/{key}_{count}.xlsx".format(path=filename, key=name, count=count))
    wbb.close()

wb.close()
